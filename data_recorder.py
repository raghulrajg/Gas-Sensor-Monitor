"""
data_recorder.py
-----------------
Handles "Record" -> "Stop -> Export to Excel" for the raw incoming sensor
data stream, plus standalone settings export/import so a calibration
profile can be reused across multiple sessions.

Design note: this is intentionally decoupled from the live-plot ring
buffers (see sensor_widget.py). The graphs only ever keep the last
N points so the UI stays smooth, but every single sample that arrives
while recording is active is appended here using openpyxl's write-only
streaming mode. Write-only mode writes rows straight through without
holding the whole sheet in memory, so a multi-hour recording will not
balloon RAM usage or slow the UI down.
"""

import datetime as _dt

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

import config


class DataRecorder:
    def __init__(self):
        self._wb = None
        self._ws = None
        self._recording = False
        self._sample_count = 0
        self._start_time = None

    @property
    def is_recording(self) -> bool:
        return self._recording

    @property
    def sample_count(self) -> int:
        return self._sample_count

    @property
    def start_time(self):
        return self._start_time

    def start(self):
        """Begin a new recording session (in-memory, streamed on write)."""
        self._wb = Workbook(write_only=True)
        self._ws = self._wb.create_sheet(title="Sensor Data")
        header = ["Timestamp", "Elapsed (s)"] + [
            f"{ch['name']} ({ch['unit']})" for ch in config.CHANNELS
        ]
        self._ws.append(header)
        self._sample_count = 0
        self._start_time = _dt.datetime.now()
        self._recording = True

    def write_sample(self, values):
        """Append one sample row. `values` is a list of NUM_CHANNELS floats."""
        if not self._recording or self._ws is None:
            return
        now = _dt.datetime.now()
        elapsed = (now - self._start_time).total_seconds()
        row = [now.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3], round(elapsed, 2)] + list(values)
        self._ws.append(row)
        self._sample_count += 1

    def stop_and_save(self, filepath: str, calibration_values=None, port_info=None):
        """Finalize the recording and write the .xlsx file to `filepath`.

        Optionally appends a second sheet with the calibration/settings
        snapshot that was active during this recording.
        """
        if self._wb is None:
            raise RuntimeError("No active recording to save.")

        if calibration_values is not None:
            self._append_settings_sheet(calibration_values, port_info)

        if not filepath.lower().endswith(".xlsx"):
            filepath += ".xlsx"

        self._wb.save(filepath)

        self._recording = False
        result = (filepath, self._sample_count)
        self._wb = None
        self._ws = None
        return result

    def discard(self):
        """Abandon the current recording without saving."""
        self._wb = None
        self._ws = None
        self._recording = False
        self._sample_count = 0

    def _append_settings_sheet(self, calibration_values, port_info):
        ws2 = self._wb.create_sheet(title="Settings Snapshot")
        ws2.append(["Recording exported", _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        if port_info:
            ws2.append(["COM Port", port_info.get("port", "")])
            ws2.append(["Baud Rate", port_info.get("baudrate", "")])
        ws2.append([])
        ws2.append(["Sensor", f"Calibration Value"])
        gas_channels = [c for c in config.CHANNELS if c["calibration"]]
        for ch, val in zip(gas_channels, calibration_values):
            ws2.append([ch["name"], val])


def export_settings_only(filepath: str, calibration_values, port_info=None):
    """Standalone export of just the current calibration/settings (no
    recorded data), used by the "Export Settings" menu action.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Settings"
    bold = Font(bold=True)

    ws.append(["Gas Sensor Monitor - Settings Export"])
    ws["A1"].font = bold
    ws.append(["Exported", _dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    if port_info:
        ws.append(["COM Port", port_info.get("port", "")])
        ws.append(["Baud Rate", port_info.get("baudrate", "")])
    ws.append([])
    ws.append(["Sensor", "Calibration Value"])
    ws["A6"].font = bold
    ws["B6"].font = bold
    gas_channels = [c for c in config.CHANNELS if c["calibration"]]
    for ch, val in zip(gas_channels, calibration_values):
        ws.append([ch["name"], val])

    if not filepath.lower().endswith(".xlsx"):
        filepath += ".xlsx"
    wb.save(filepath)
    return filepath


def import_settings(filepath: str) -> dict:
    """Read a previously exported settings file (either from
    `export_settings_only` or the "Settings Snapshot" sheet written
    alongside a recording) and return {sensor_name: calibration_value}.

    Tolerant of either sheet name ("Settings" or "Settings Snapshot") and
    simply scans for the "Sensor" / "Calibration Value" header row, then
    reads name/value pairs until it hits a blank row.
    """
    wb = load_workbook(filepath, data_only=True)

    candidate_sheets = [s for s in wb.sheetnames if "Settings" in s] or wb.sheetnames
    for sheet_name in candidate_sheets:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        header_row_idx = None
        for i, row in enumerate(rows):
            if row and row[0] == "Sensor":
                header_row_idx = i
                break
        if header_row_idx is None:
            continue

        result = {}
        for row in rows[header_row_idx + 1:]:
            if not row or row[0] in (None, ""):
                break
            name = row[0]
            try:
                value = float(row[1])
            except (TypeError, ValueError, IndexError):
                continue
            result[name] = value
        if result:
            return result

    raise ValueError(
        "Couldn't find a calibration settings table in this file. "
        "Expected a sheet with a 'Sensor' / 'Calibration Value' header row."
    )
